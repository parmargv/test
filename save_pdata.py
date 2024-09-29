import os
import pandas as pd
import streamlit as st
import openpyxl
from datetime import datetime

def save_pro(data):
    absolute_path = os.path.dirname(__file__)
    file_path = os.path.join(absolute_path, 'p_data.xlsx')
    wb = openpyxl.load_workbook(file_path)
    sh1 = wb['data']

    headers = ['DATE', 'FII_CALL','FII_PUT','FII_NET','Date','PRO_CALL','PRO_PUT','PRO_NET']
    for col_num, header in enumerate(headers, start=1):
        sh1.cell(row=1, column=col_num, value=header)

    # Find the next available row starting from row 2

    next_row = 2
    while sh1.cell(row=next_row, column=1).value is not None:
        next_row += 1

    # Write data to the next available row
    for col_num, value in enumerate(data, start=1):
        sh1.cell(row=next_row, column=col_num, value=value)
    # Save the workbook
    wb.save('p_data.xlsx')
    wb.close()
def read_data():
    absolute_path = os.path.dirname(__file__)
    file_path = os.path.join(absolute_path, 'p_data.xlsx')
    wb = openpyxl.load_workbook(file_path)
    sh1 = wb['data']
    data = []

    # Iterate over the rows in the worksheet
    for row in sh1.iter_rows(min_row=2,max_col=4, values_only=True):
        if not all(cell is None for cell in row):
            data.append(row)

    # Convert the data to a DataFrame
    df1 = pd.DataFrame(data, columns=['DATE', 'FII_CALL', 'FII_PUT', 'FII_NET'])
    # df['Date'] = pd.to_datetime(df['Date'], format='%d-%m-%Y')
    # # Sort the DataFrame by the 'Date' column in ascending order
    df1= df1.sort_values(by='DATE')

    # Reset the index to reflect the sorted order
    df1.reset_index(drop=True, inplace=True)

    for row in sh1.iter_rows(min_row=2,min_col=5, max_col=8, values_only=True):
        if not all(cell is None for cell in row):
            data.append(row)

    # Convert the data to a DataFrame
    df2 = pd.DataFrame(data, columns=['Date', 'PRO_CALL', 'PRO_PUT', 'PRO_NET'])
    # df['Date'] = pd.to_datetime(df['Date'], format='%d-%m-%Y')
    # # Sort the DataFrame by the 'Date' column in ascending order
    df2 = df2.sort_values(by='Date')

    # Reset the index to reflect the sorted order
    df2.reset_index(drop=True, inplace=True)
    # Print the DataFrame
    return df1,df2
def clear_data():
    absolute_path = os.path.dirname(__file__)
    file_path = os.path.join(absolute_path, 'p_data.xlsx')
    wb = openpyxl.load_workbook(file_path)
    sh1 = wb['data']
    start_row = 2
    end_row = 6
    start_col = 1
    end_col = 8

    # Clear the range by setting each cell's value to None
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = sh1.cell(row=row, column=col)
            cell.value = None

    # Save the workbook
    wb.save('p_data.xlsx')
    wb.close()